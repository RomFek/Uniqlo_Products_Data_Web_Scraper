from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, colors

class Uniqlo_Scraper():
    
        def __init__(self):   
            self._file_name = 'items_data.xml'
            self._uniqlo_url = "https://www.uniqlo.com" #URL of the tickets page on Filmstaden website
            self._headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36'} #Fake a user-agent to prevent the website from blocking our requests

            #https://www.uniqlo.com/se/en/home
            
        @property 
        def uniqlo_url(self):
            return self._uniqlo_url
        
        @property 
        def headers(self):
            return self._headers
    
        def check_uniqlo_connection(self):
            error_message = "[ERROR] Could not establish connection to Uniqlo with the provided URL."
            success_message = "Connection to Uniqlo established successfully."
            print(requests.get(self.uniqlo_url, headers=self.headers).status_code)

            try:
                print("Checking connection to Uniqlo...")    
                if requests.get(self.uniqlo_url, headers=self.headers).status_code == 200:
                    print(success_message)
                    return True
                else:
                    print(error_message)
                    return False
            except:
                print(error_message)
                return False
            
        #Get product category links
        def get_all_links(self):
            print("Retrieving links to product categories")
            try:
                uniqlo_request = requests.get(self.uniqlo_url, headers=self.headers)
                u_soup = BeautifulSoup(uniqlo_request.text, 'html.parser')
                links = u_soup.findAll('a', {'class':'mainnav__link'})
                link_list = []
                for link in links:
                    link_url = link.get('href')
                    link_list.append(link_url)
                print("Links for product categories retrieved")
                return link_list
            except:
                print("[ERROR] An error occurred while retrieving the category links")
                return None
        
        #Get  links for all subcategories in each category
        def get_all_subcategory_links(self):
            try:
                print("Retrieving sub-category links for each category")
                category_links = self.get_all_links()
                all_data = []
                for subcategory_link in category_links:
                    category_data = {}
                    uniqlo_request = requests.get(subcategory_link, headers=self.headers)
                    u_soup = BeautifulSoup(uniqlo_request.text, 'html.parser')
                    link_sections = u_soup.findAll('a', {'class':'l2Hero__link'})
                    subcategory_link_list = []
                    category_name = None
                    for section in link_sections:
                        #section_soup = BeautifulSoup(str(section), 'html.parser') #Feed the html to BeautifulSoap
                        #link = section_soup.find('a', {'href':'originalTitle'})
                        category_name = section.get('data-category')
                        link = section.get('href')
                        if link != 'None':
                            if link not in subcategory_link_list: #Avoid duplicate links in case if such exist
                                subcategory_link_list.append(link) 
                    category_data['CategoryName'] = category_name
                    category_data['SubCategoryLinks'] = subcategory_link_list
                    print(subcategory_link_list)
                    if category_name != None:
                        all_data.append(category_data)
                        print("Sub-category links for category '" + category_name +  "' retrieved successfully...")
                    #print(category_name)
                    #print(subcategory_link_list)
                return all_data
            except:
                print("[ERROR] An error occurred while trying to retrieve sub-category links")
                return None
            
        def get_product_links(self):
            print("Retrieving product links from each category's sub-category...")
            products = []
            links = self.get_all_subcategory_links()
            #try:
            for category_data in links:
                print(category_data['CategoryName'])
                for link in category_data['SubCategoryLinks']:
                    #Exclude the links to lookbook pages
                    if 'lookbook' not in link: 
                        print(link)
                        uniqlo_request = requests.get(link, headers=self.headers)
                        u_soup = BeautifulSoup(uniqlo_request.text, 'html.parser')
                        product_segments = u_soup.findAll('a', {'class':'productTile__link name-link fontUniqloProRegular'})
                        for product in product_segments:
                            product_data = {}
                            product_link = product.get('href')
                            product_name = product.get('title')
                            product_page_url = self.uniqlo_url + product_link
                            request = requests.get(product_page_url, headers=self.headers)
                            pr_u_soup = BeautifulSoup(request.text, 'html.parser')
                            product_segments = pr_u_soup.findAll('p', {'class':'deliverySection__text'})
                            try:
                                product_material = product_segments[0].text
                            except: 
                                product_material = None
                            try:
                                product_care = product_segments[1].text
                            except:
                                product_care = None
                            try:
                                product_sku = pr_u_soup.findAll('span', {'itemprop':'productID'})[0].text
                            except:
                                product_sku = None
                            #print(product_name + ' url: ' + product_link + " Material: " + product_material + " Care Instructions: " + product_care)
                            product_data['Category'] = category_data['CategoryName']
                            product_data['ProductName'] = product_name
                            product_data['ProductLink'] = product_page_url
                            print(product_page_url)
                            product_data['ProductMaterial'] = product_material
                            product_data['CareInstructions'] = product_care
                            product_data['ProductSKU'] = product_sku

                            print(product_data)
                            products.append(product_data)
                #print(products)
            print("Product data retrieved...")
            return products
            #except:
            #    print("[ERROR] An error occurred while retrieving product data.")
            #    return None
         
        def generate_excel(self):
            products = self.get_product_links()
            print("Generating excel...")
            if products is not None:
                try:
                    wb = Workbook()
                    wb.remove(wb['Sheet']) #Delete default sheet
                    women_sheet = wb.create_sheet("Sheet_A", 0)
                    women_sheet.title = "Women"
                    men_sheet = wb.create_sheet("Sheet_B", 1)
                    men_sheet.title = "Men"
                    kids_sheet = wb.create_sheet("Sheet_C", 2)
                    kids_sheet.title = "Kids"
                    baby_sheet = wb.create_sheet("Sheet_D", 3)
                    baby_sheet.title = "Baby"   
                    #Set the title of the column header of each sheet
                    women_sheet['A1'] = men_sheet['A1'] = kids_sheet['A1'] = baby_sheet['A1'] ='Category' 
                    women_sheet['B1'] = men_sheet['B1'] = kids_sheet['B1'] = baby_sheet['B1'] ='ProductName' 
                    women_sheet['C1'] = men_sheet['C1'] = kids_sheet['C1'] = baby_sheet['C1'] ='ProductSKU' 
                    women_sheet['D1'] = men_sheet['D1'] = kids_sheet['D1'] = baby_sheet['D1'] ='ProductMaterial' 
                    women_sheet['E1'] = men_sheet['E1'] = kids_sheet['E1'] = baby_sheet['E1'] ='CareInstructions' 
                    women_sheet['F1'] = men_sheet['F1'] = kids_sheet['F1'] = baby_sheet['F1'] ='ProductLink' 
                    women_sheet['G1'] = men_sheet['G1'] = kids_sheet['G1'] = baby_sheet['G1'] ='ProductLink' 
                    #Set the font color and style of the column header of each sheet
                    font_color = Font(size=11, color = colors.WHITE, bold=True, italic=False) 
                    women_sheet['A1'].font = women_sheet['B1'].font = women_sheet['C1'].font = women_sheet['D1'].font = women_sheet['E1'].font = women_sheet['F1'].font = women_sheet['G1'].font= font_color
                    men_sheet['A1'].font = men_sheet['B1'].font = men_sheet['C1'].font = men_sheet['D1'].font = men_sheet['E1'].font = men_sheet['F1'].font = men_sheet['G1'].font= font_color
                    kids_sheet['A1'].font = kids_sheet['B1'].font = kids_sheet['C1'].font = kids_sheet['D1'].font = kids_sheet['E1'].font = kids_sheet['F1'].font = kids_sheet['G1'].font= font_color
                    baby_sheet['A1'].font = baby_sheet['B1'].font = baby_sheet['C1'].font = baby_sheet['D1'].font = baby_sheet['E1'].font = baby_sheet['F1'].font = baby_sheet['G1'].font= font_color
                    #Set the width of the column header cells of each sheet
                    women_sheet.column_dimensions['A'].width = men_sheet.column_dimensions['A'].width  = kids_sheet.column_dimensions['A'].width = baby_sheet.column_dimensions['A'].width  = 15
                    women_sheet.column_dimensions['B'].width = men_sheet.column_dimensions['B'].width  = kids_sheet.column_dimensions['B'].width = baby_sheet.column_dimensions['B'].width  = 70
                    women_sheet.column_dimensions['C'].width = men_sheet.column_dimensions['C'].width  = kids_sheet.column_dimensions['C'].width = baby_sheet.column_dimensions['C'].width  = 15
                    women_sheet.column_dimensions['D'].width = men_sheet.column_dimensions['D'].width  = kids_sheet.column_dimensions['D'].width = baby_sheet.column_dimensions['D'].width  = 70
                    women_sheet.column_dimensions['E'].width = men_sheet.column_dimensions['E'].width  = kids_sheet.column_dimensions['E'].width = baby_sheet.column_dimensions['E'].width  = 70
                    women_sheet.column_dimensions['G'].width = men_sheet.column_dimensions['G'].width  = kids_sheet.column_dimensions['G'].width = baby_sheet.column_dimensions['G'].width  = 70
                    #Hide the product link column
                    women_sheet.column_dimensions['F'].hidden = men_sheet.column_dimensions['F'].hidden  = kids_sheet.column_dimensions['F'].hidden = baby_sheet.column_dimensions['F'].hidden  = True

                    #Set the cell color of the column header of each sheet
                    cell_fill = PatternFill(fill_type='solid', start_color='231750', end_color='231750')
                    women_sheet['A1'].fill = women_sheet['B1'].fill = women_sheet['C1'].fill = women_sheet['D1'].fill = women_sheet['E1'].fill = women_sheet['F1'].fill = women_sheet['G1'].fill = cell_fill
                    men_sheet['A1'].fill = men_sheet['B1'].fill = men_sheet['C1'].fill = men_sheet['D1'].fill = men_sheet['E1'].fill = men_sheet['F1'].fill = men_sheet['G1'].fill= cell_fill
                    kids_sheet['A1'].fill = kids_sheet['B1'].fill = kids_sheet['C1'].fill = kids_sheet['D1'].fill = kids_sheet['E1'].fill = kids_sheet['F1'].fill = kids_sheet['G1'].fill= cell_fill
                    baby_sheet['A1'].fill = baby_sheet['B1'].fill = baby_sheet['C1'].fill = baby_sheet['D1'].fill = baby_sheet['E1'].fill = baby_sheet['F1'].fill = baby_sheet['G1'].fill = cell_fill
                    
                    #Turn the columns in the first row into column filters on each sheet
                    women_sheet.auto_filter.ref = 'A1:G1'
                    men_sheet.auto_filter.ref = 'A1:G1'
                    kids_sheet.auto_filter.ref = 'A1:G1'
                    baby_sheet.auto_filter.ref = 'A1:G1'

                    counterM = counterW = counterK = counterB = 1 #Start with row 2 because 1st row is for headers
                    for product in products:
                        if product['Category'] == 'Women':
                            counterW += 1
                            women_sheet['A' + str(counterW)] = product['Category']
                            women_sheet['B' + str(counterW)] = product['ProductName']
                            women_sheet['C' + str(counterW)] = product['ProductSKU']
                            women_sheet['D' + str(counterW)] = product['ProductMaterial']
                            women_sheet['E' + str(counterW)] = product['CareInstructions']
                            women_sheet['F' + str(counterW)] = product['ProductLink']
                            women_sheet['G' + str(counterW)] = "=Hyperlink(F" + str(counterW) + ")"
                        elif product['Category'] == 'Men':
                            counterM += 1
                            men_sheet['A' + str(counterM)] = product['Category']
                            men_sheet['B' + str(counterM)] = product['ProductName']
                            men_sheet['C' + str(counterM)] = product['ProductSKU']
                            men_sheet['D' + str(counterM)] = product['ProductMaterial']
                            men_sheet['E' + str(counterM)] = product['CareInstructions']
                            men_sheet['F' + str(counterM)] = product['ProductLink']
                            men_sheet['G' + str(counterM)] = "=Hyperlink(F" + str(counterM) + ")"
                        elif product['Category'] == 'Kids':
                            counterK += 1
                            kids_sheet['A' + str(counterK)] = product['Category']
                            kids_sheet['B' + str(counterK)] = product['ProductName']
                            kids_sheet['C' + str(counterK)] = product['ProductSKU']
                            kids_sheet['D' + str(counterK)] = product['ProductMaterial']
                            kids_sheet['E' + str(counterK)] = product['CareInstructions']
                            kids_sheet['F' + str(counterK)] = product['ProductLink']
                            kids_sheet['G' + str(counterK)] = "=Hyperlink(F" + str(counterK) + ")"
                        elif product['Category'] == 'Baby':
                            counterB += 1
                            baby_sheet['A' + str(counterB)] = product['Category']
                            baby_sheet['B' + str(counterB)] = product['ProductName']
                            baby_sheet['C' + str(counterB)] = product['ProductSKU']
                            baby_sheet['D' + str(counterB)] = product['ProductMaterial']
                            baby_sheet['E' + str(counterB)] = product['CareInstructions']
                            baby_sheet['F' + str(counterB)] = product['ProductLink']
                            baby_sheet['G' + str(counterB)] = "=Hyperlink(F" + str(counterB) + ")"

                        else:
                            print('Category not recognized.')
                    #ws1['B1'] = "Material"
                    #ws1['C1'] = "Care Instructions"
                    wb.save(filename="product_list.xlsx")
                    print("Excel generated successfully!")
                except:
                    print("[ERROR] An error occurred while generating the excel file.")
               
sc = Uniqlo_Scraper()
sc.generate_excel()
